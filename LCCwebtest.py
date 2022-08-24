
import LCC
import requests
# 爬取ig關鍵字圖片
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
from bs4 import BeautifulSoup
import os
import urllib.request
from selenium.webdriver.chrome.options import Options
import openpyxl

PATH = "G:/This_is_my_qt_project/chromedriver.exe"
driver = webdriver.Chrome(PATH)
driver.get("https://paperswithcode.com/")
time.sleep(3)






search = driver.find_element_by_xpath('/html/body/nav/button')
search.click()
search = driver.find_element_by_xpath('//*[@id="id_global_search_input"]')
search.send_keys("pose estimation")
search.click()
time.sleep(3)
search = driver.find_element_by_xpath('//*[@id="id_global_search_form"]/button')
search.click()
time.sleep(3)
for i in range(1,2):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)

soup = BeautifulSoup(driver.page_source, 'html.parser')
#print(soup)
hh_total = []
hh_paper = []
hh_code = []




for link in soup.select('.entity a'):
    hh_total.append('https://paperswithcode.com'+str(link.get('href')))
#print(hh_total)
for i in range(len(hh_total)):
    if '#code' in hh_total[i]:
        hh_code.append(hh_total[i])
    else:
        hh_paper.append(hh_total[i])
#print(hh_paper)
# PATH = "G:/This_is_my_qt_project/chromedriver.exe"
#driver2  = webdriver.Chrome(PATH)
hh_paper = set(hh_paper)
hh_paper = list(hh_paper)




hh_con = []
wb = openpyxl.Workbook() 
wb.save('paper.xlsx')
for url in hh_paper:
    # PATH = "G:/This_is_my_qt_project/chromedriver.exe"
    #driver  = webdriver.Chrome(PATH)
    driver.get(url)
    time.sleep(1)
    soup2 = BeautifulSoup(driver.page_source, 'html.parser')
    hh_pdf = []
    for title in soup2.select('.paper-title h1'):
        titlee = title.getText().replace(' ','')
        hh_pdf.append(titlee.replace('\n',''))
        
        #print(hh_pdf)
    for link in soup2.select('.item-conference-link a'):
        linkk = link.getText().replace(' ','')
        hh_pdf.append(linkk.replace('\n',''))
        
        #print(hh_pdf)
    
    for link in soup2.select('.row a'):
        if 'https' in str(link.get('href')):
            if '.pdf' in str(link.get('href')):
                print(link.get('href'))
                
                hh_pdf.append(link.get('href'))
                hh_con.append(hh_pdf)
                time.sleep(1)

            #print(hh_pdf)
        # else :
        # hh_pdf.append(link.get('href'))
        # hh_con.append(hh_pdf)
        # time.sleep(3)
        #print(hh_con)
    #print(hh_con)    
#print(hh_con)


wb  = openpyxl.load_workbook('paper.xlsx',data_only=True)
s1 = wb.create_sheet('paper')
for i in hh_con:
    s1.append(i)
wb.save('paper.xlsx')
