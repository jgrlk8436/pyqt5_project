from PyQt5 import QtWidgets, QtGui, QtCore
import LCC
import requests
# 爬取ig關鍵字圖片
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
from bs4 import BeautifulSoup
import os
import urllib.request
from selenium.webdriver.chrome.options import Options
import openpyxl


class MainWindow_controller_LCC(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__() # in python3, super(Class, self).xxx = super().xxx
        self.ui = LCC.Ui_MainWindow()
        self.course = ['pose estimation','segmentation','tracking','prunning','voice'
        ]
        self.ui.setupUi(self)
        self.setup_c()

    def setup_c(self):
        self.clicked_counter = 0
        self.ui.comboBox.addItems(self.course)
        self.ui.pushButton.clicked.connect(self.download)
        # self.ui.Data_Analysis_btn.clicked.connect(self.data_analysisClicked)
        # self.ui.Computer_Vision_btn.clicked.connect(self.computer_visionClicked)

    def download(self):
        #print(f'{self.ui.comboBox.currentText()}')
        string = f'{self.ui.comboBox.currentText()}'

        
        PATH = "chromedriver.exe"
        driver = webdriver.Chrome(PATH)

        driver.get("https://paperswithcode.com/")
        time.sleep(3)

        
        search = driver.find_element_by_xpath('/html/body/nav/button')
        search.click()
        search = driver.find_element_by_xpath('//*[@id="id_global_search_input"]')
        search.send_keys(string)
        search.click()
        time.sleep(3)
        search = driver.find_element_by_xpath('//*[@id="id_global_search_form"]/button')
        search.click()
        time.sleep(3)
        for i in range(1,2):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        #print(soup)
        hh_total = []
        hh_paper = []
        hh_code = []

        
                
        for link in soup.select('.entity a'):
            hh_total.append('https://paperswithcode.com'+str(link.get('href')))
        #print(hh_total)
        for i in range(len(hh_total)):
            if '#code' in hh_total[i]:
                hh_code.append(hh_total[i])
            else:
                hh_paper.append(hh_total[i])
        #print(hh_paper)
        # PATH = "G:/This_is_my_qt_project/chromedriver.exe"
        #driver2  = webdriver.Chrome(PATH)
        hh_paper = set(hh_paper)
        hh_paper = list(hh_paper)




        hh_con = []
        wb = openpyxl.Workbook() 
        wb.save('paper.xlsx')
        for url in hh_paper:
            driver.get(url)
            time.sleep(1)
            soup2 = BeautifulSoup(driver.page_source, 'html.parser')
            hh_pdf = []
            for title in soup2.select('.paper-title h1'):
                titlee = title.getText().replace(' ','')
                hh_pdf.append(titlee.replace('\n',''))
            for link in soup2.select('.item-conference-link a'):
                linkk = link.getText().replace(' ','')
                hh_pdf.append(linkk.replace('\n',''))
            for link in soup2.select('.row a'):
                if 'https' in str(link.get('href')):
                    if '.pdf' in str(link.get('href')):
                        print(link.get('href'))
                        hh_pdf.append(link.get('href'))
                        hh_con.append(hh_pdf)
                        time.sleep(1)
        
        wb  = openpyxl.load_workbook('paper.xlsx',data_only=True)
        s1 = wb.create_sheet('paper')
        for i in hh_con:
            s1.append(i)
        wb.save('paper.xlsx')